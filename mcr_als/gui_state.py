"""Display-independent state, parsing, and file I/O for the desktop GUI."""

from __future__ import annotations

from dataclasses import asdict, dataclass, field, fields
from pathlib import Path
from typing import Any, Mapping, Sequence

import numpy as np
from numpy.typing import ArrayLike, NDArray
from scipy.io import loadmat  # type: ignore[import-untyped]

from .helpers import efa, pure
from .options import (
    ClosureBlock,
    ClosureCondition,
    ClosureOptions,
    CorrelationOptions,
    KineticModelOptions,
    KineticOptions,
    MCRALSOptions,
    MultiExperimentOptions,
    NonnegativityOptions,
    TrilinearityOptions,
    TuckerModeOptions,
    TuckerOptions,
    UnimodalityOptions,
    ValueConstraint,
    WeightedOptions,
)

FloatArray = NDArray[np.float64]


@dataclass(slots=True)
class KineticModelInput:
    """Text/file fields used by the kinetic-model editor.

    Component mappings intentionally use the legacy GUI convention: zero
    ignores an ALS component and positive values are one-based species IDs.
    """

    name: str = "Kinetic model"
    reaction_orders: str = "1 0"
    stoichiometry: str = "-1; 1"
    initial_rate_constants: str = "0.1"
    initial_concentrations: str = "1 0"
    time: str = "0 1 2 3 4 5"
    component_mapping: str = "1 2"
    colored_mask: str = "1 1"
    experiment_mask: str = ""


@dataclass(slots=True)
class GUIState:
    """Serializable state for the main MCR-ALS desktop window."""

    data_path: str = ""
    initial_path: str = ""
    initial_method: str = "file"
    components: int = 2
    pure_noise_percent: float = 5.0
    row_lengths: str = ""
    column_lengths: str = ""
    presence: str = ""

    max_iterations: int = 50
    tolerance: float = 0.1
    divergence_limit: int = 20
    normalization: str = "none"

    nonnegative_c_enabled: bool = True
    nonnegative_c_algorithm: str = "truncate"
    nonnegative_c_mask: str = ""
    nonnegative_s_enabled: bool = True
    nonnegative_s_algorithm: str = "truncate"
    nonnegative_s_mask: str = ""
    unimodality_c_enabled: bool = False
    unimodality_c_tolerance: float = 1.1
    unimodality_c_mode: int = 1
    unimodality_c_mask: str = ""
    unimodality_s_enabled: bool = False
    unimodality_s_tolerance: float = 1.1
    unimodality_s_mode: int = 1
    unimodality_s_mask: str = ""

    closure_enabled: bool = False
    closure_mode: str = "concentration"
    closure_kind: str = "equality"
    closure_target: str = "1"
    closure_components: str = ""
    closure_second_enabled: bool = False
    closure_second_kind: str = "equality"
    closure_second_target: str = "1"
    closure_second_components: str = ""
    concentration_values_path: str = ""
    concentration_values_kind: str = "equal"
    spectral_values_path: str = ""
    spectral_values_kind: str = "equal"

    correlation_enabled: bool = False
    correlation_reference_path: str = ""
    correlation_component_mask: str = ""
    correlation_model: str = "global"
    correlation_matrix_effect: bool = False
    correlation_normalize_spectra: bool = False

    trilinearity_enabled: bool = False
    trilinearity_direction: str = "concentration"
    trilinearity_shape: int = 1
    trilinearity_component_mask: str = ""
    quadrilinear_dimensions: str = ""

    tucker_enabled: bool = False
    tucker_n_matrices: int = 0
    tucker_mode1_groups: str = ""
    tucker_mode3_groups: str = ""

    weighted_enabled: bool = False
    standard_deviations_path: str = ""
    weighted_convergence_limit: float = 1.0e-10
    weighted_max_iterations: int = 200_000

    kinetic_enabled: bool = False
    kinetic_models: list[KineticModelInput] = field(default_factory=list)


def parse_matrix_text(text: str, *, name: str = "matrix") -> FloatArray:
    """Parse whitespace/comma-separated values with newlines or semicolons."""
    cleaned = text.strip()
    if cleaned.startswith("[") and cleaned.endswith("]"):
        cleaned = cleaned[1:-1].strip()
    if not cleaned:
        raise ValueError(f"{name} is empty")
    rows = [row.strip() for row in cleaned.replace(";", "\n").splitlines()]
    parsed: list[list[float]] = []
    for row in rows:
        if not row:
            continue
        tokens = row.replace(",", " ").split()
        try:
            parsed.append([float(token) for token in tokens])
        except ValueError as exc:
            raise ValueError(f"{name} contains a non-numeric value") from exc
    if not parsed or not parsed[0]:
        raise ValueError(f"{name} is empty")
    width = len(parsed[0])
    if any(len(row) != width for row in parsed):
        raise ValueError(f"{name} rows must have equal lengths")
    result = np.asarray(parsed, dtype=np.float64)
    if not np.all(np.isfinite(result) | np.isnan(result)):
        raise ValueError(f"{name} contains an invalid numeric value")
    return result


def parse_float_vector(text: str, *, name: str = "values") -> FloatArray:
    return parse_matrix_text(text, name=name).reshape(-1)


def parse_int_vector(text: str, *, name: str = "values") -> NDArray[np.intp]:
    values = parse_float_vector(text, name=name)
    rounded = np.rint(values)
    if not np.array_equal(values, rounded):
        raise ValueError(f"{name} must contain integers")
    return rounded.astype(np.intp)


def parse_optional_mask(
    text: str,
    size: int,
    *,
    name: str,
) -> NDArray[np.bool_] | None:
    if not text.strip():
        return None
    values = parse_int_vector(text, name=name)
    if values.size != size or np.any((values != 0) & (values != 1)):
        raise ValueError(f"{name} must contain {size} zero/one entries")
    return values.astype(bool)


def parse_lengths(text: str, *, name: str) -> tuple[int, ...] | None:
    if not text.strip():
        return None
    values = parse_int_vector(text, name=name)
    if np.any(values <= 0):
        raise ValueError(f"{name} must be positive")
    return tuple(int(value) for value in values)


def _source_parts(source: str | Path) -> tuple[Path, str | None]:
    text = str(source).strip()
    if "::" in text:
        path_text, variable = text.rsplit("::", 1)
        return Path(path_text), variable or None
    return Path(text), None


def _pick_array(
    arrays: Mapping[str, Any],
    variable: str | None,
    *,
    source: Path,
) -> FloatArray:
    if variable is not None:
        if variable not in arrays:
            raise ValueError(f"{source.name} has no array named {variable!r}")
        candidate = arrays[variable]
    else:
        candidates: list[tuple[str, np.ndarray]] = []
        for key, raw in arrays.items():
            if key.startswith("__"):
                continue
            array = np.asarray(raw)
            if np.issubdtype(array.dtype, np.number) and array.ndim >= 1:
                candidates.append((key, array))
        if not candidates:
            raise ValueError(f"{source.name} contains no numeric arrays")
        candidate = max(candidates, key=lambda item: item[1].size)[1]
    array = np.asarray(candidate)
    if np.iscomplexobj(array) and np.any(np.imag(array) != 0.0):
        raise ValueError("complex-valued matrices are not supported")
    return np.asarray(np.real(array), dtype=np.float64)


def load_numeric_array(source: str | Path) -> FloatArray:
    """Load an array from text, NumPy, MATLAB, or Excel data.

    ``file.mat::variable`` and ``file.npz::key`` select a named array. Without
    a name, the largest numeric array is selected.
    """
    path, variable = _source_parts(source)
    if not path.is_file():
        raise FileNotFoundError(path)
    suffix = path.suffix.lower()
    if suffix == ".npy":
        array = np.load(path, allow_pickle=False)
    elif suffix == ".npz":
        with np.load(path, allow_pickle=False) as archive:
            array = _pick_array(dict(archive.items()), variable, source=path)
    elif suffix == ".mat":
        array = _pick_array(loadmat(path), variable, source=path)
    elif suffix in (".csv", ".tsv", ".txt", ".dat"):
        delimiter = "," if suffix == ".csv" else "\t" if suffix == ".tsv" else None
        try:
            array = np.loadtxt(path, delimiter=delimiter, dtype=np.float64)
        except ValueError:
            if delimiter is not None:
                raise
            array = np.loadtxt(path, delimiter=",", dtype=np.float64)
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook  # type: ignore[import-untyped]
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to load .xlsx files") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = [
            [np.nan if value is None else float(value) for value in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
        workbook.close()
        array = np.asarray(rows, dtype=np.float64)
    else:
        raise ValueError(
            "supported inputs are .csv, .tsv, .txt, .dat, .npy, .npz, .mat, and .xlsx"
        )
    result = np.asarray(array, dtype=np.float64)
    if result.ndim == 0 or result.ndim > 2:
        raise ValueError("loaded numeric arrays must be one- or two-dimensional")
    return result


def load_matrix(source: str | Path) -> FloatArray:
    array = load_numeric_array(source)
    if array.ndim == 1:
        array = array.reshape(1, -1)
    if array.size == 0:
        raise ValueError("matrix is empty")
    return np.asarray(array, dtype=np.float64)


def array_from_source(text: str, *, name: str) -> FloatArray:
    source = text.strip()
    if not source:
        raise ValueError(f"{name} is empty")
    try:
        path, _ = _source_parts(source)
        if path.is_file():
            return load_numeric_array(source)
    except OSError:
        pass
    return parse_matrix_text(source, name=name)


def format_matrix(array: ArrayLike) -> str:
    values = np.asarray(array)
    if values.ndim == 1:
        values = values[None, :]
    return ";\n".join(
        " ".join(f"{float(value):.12g}" for value in row) for row in values
    )


def kinetic_model_from_input(model: KineticModelInput) -> KineticModelOptions:
    orders = array_from_source(model.reaction_orders, name="reaction orders")
    stoichiometry = array_from_source(model.stoichiometry, name="stoichiometry")
    rates = array_from_source(
        model.initial_rate_constants, name="initial rate constants"
    ).reshape(-1)
    initial = array_from_source(
        model.initial_concentrations, name="initial concentrations"
    )
    time = array_from_source(model.time, name="kinetic time")
    legacy_mapping = parse_int_vector(model.component_mapping, name="component mapping")
    if np.any(legacy_mapping < 0):
        raise ValueError("GUI component mappings use 0=ignore and positive species IDs")
    mapping = np.where(legacy_mapping == 0, -1, legacy_mapping - 1)
    colored = parse_int_vector(model.colored_mask, name="colored mask")
    if np.any((colored != 0) & (colored != 1)):
        raise ValueError("colored mask must contain only zero and one")
    experiment_mask: NDArray[np.intp] | None = None
    if model.experiment_mask.strip():
        experiment_mask = parse_int_vector(
            model.experiment_mask, name="experiment mask"
        )
        if np.any((experiment_mask != 0) & (experiment_mask != 1)):
            raise ValueError("experiment mask must contain only zero and one")
    return KineticModelOptions(
        reaction_orders=orders,
        stoichiometry=stoichiometry,
        initial_rate_constants=rates,
        initial_concentrations=initial,
        time=time,
        component_mapping=mapping,
        colored_mask=colored,
        experiment_mask=experiment_mask,
        choose=legacy_mapping,
        name=model.name.strip() or "Kinetic model",
    )


def _target_from_text(text: str, *, name: str) -> float | FloatArray:
    values = array_from_source(text, name=name)
    if values.size == 1:
        return float(values.reshape(-1)[0])
    return values


def _value_constraint(path: str, kind: str) -> ValueConstraint | None:
    if not path.strip():
        return None
    return ValueConstraint(values=load_matrix(path), kind=kind)  # type: ignore[arg-type]


def build_options(
    state: GUIState,
    data_shape: tuple[int, int],
    n_components: int,
) -> MCRALSOptions:
    """Translate GUI state into the public numerical API dataclasses."""
    row_lengths = parse_lengths(state.row_lengths, name="row lengths")
    column_lengths = parse_lengths(state.column_lengths, name="column lengths")
    if row_lengths is not None and sum(row_lengths) != data_shape[0]:
        raise ValueError("row lengths must sum to the number of data rows")
    if column_lengths is not None and sum(column_lengths) != data_shape[1]:
        raise ValueError("column lengths must sum to the number of data columns")
    presence: NDArray[np.bool_] | None = None
    if state.presence.strip():
        raw_presence = parse_matrix_text(state.presence, name="component presence")
        expected_rows = 1 if row_lengths is None else len(row_lengths)
        if raw_presence.shape != (expected_rows, n_components):
            raise ValueError("component presence must be row-blocks by ALS components")
        if np.any((raw_presence != 0) & (raw_presence != 1)):
            raise ValueError("component presence must contain only zero and one")
        presence = raw_presence.astype(bool)
    multi = MultiExperimentOptions.from_lengths(
        row_lengths=row_lengths,
        column_lengths=column_lengths,
        presence=presence,
    )

    closure_blocks: tuple[ClosureBlock, ...] = ()
    if state.closure_enabled:
        components = parse_optional_mask(
            state.closure_components,
            n_components,
            name="closure components",
        )
        if components is None:
            components = np.ones(n_components, dtype=bool)
        first = ClosureCondition(
            components=components,
            kind=state.closure_kind,  # type: ignore[arg-type]
            target=_target_from_text(state.closure_target, name="closure target"),
        )
        second: ClosureCondition | None = None
        if state.closure_second_enabled:
            second_components = parse_optional_mask(
                state.closure_second_components,
                n_components,
                name="second closure components",
            )
            if second_components is None:
                raise ValueError("second closure components are required")
            second = ClosureCondition(
                components=second_components,
                kind=state.closure_second_kind,  # type: ignore[arg-type]
                target=_target_from_text(
                    state.closure_second_target, name="second closure target"
                ),
            )
        block_count = (
            (1 if row_lengths is None else len(row_lengths))
            if state.closure_mode == "concentration"
            else (1 if column_lengths is None else len(column_lengths))
        )
        closure_blocks = tuple(
            ClosureBlock(first=first, second=second) for _ in range(block_count)
        )

    correlation: CorrelationOptions | None = None
    if state.correlation_enabled:
        if not state.correlation_reference_path.strip():
            raise ValueError("a correlation reference file is required")
        correlation = CorrelationOptions(
            reference=load_matrix(state.correlation_reference_path),
            component_mask=parse_optional_mask(
                state.correlation_component_mask,
                n_components,
                name="correlation component mask",
            ),
            model=state.correlation_model,  # type: ignore[arg-type]
            matrix_effect=state.correlation_matrix_effect,
            normalize_spectra=state.correlation_normalize_spectra,
        )

    quadrilinear: tuple[int, int, int] | None = None
    if state.quadrilinear_dimensions.strip():
        dimensions = parse_int_vector(
            state.quadrilinear_dimensions, name="quadrilinear dimensions"
        )
        if dimensions.size != 3 or np.any(dimensions <= 0):
            raise ValueError("quadrilinear dimensions require three positive integers")
        quadrilinear = tuple(int(value) for value in dimensions)  # type: ignore[assignment]

    tucker_modes: list[TuckerModeOptions] = []
    for mode, raw_groups in (
        (1, state.tucker_mode1_groups),
        (3, state.tucker_mode3_groups),
    ):
        if raw_groups.strip():
            legacy_groups = parse_int_vector(
                raw_groups, name=f"Tucker mode {mode} groups"
            )
            if legacy_groups.size != n_components:
                raise ValueError("Tucker groups must match ALS components")
            groups = np.where(legacy_groups > 0, legacy_groups - 1, -1)
            tucker_modes.append(TuckerModeOptions(mode=mode, groups=groups))  # type: ignore[arg-type]
    if state.tucker_enabled and not tucker_modes:
        raise ValueError("enabled Tucker constraints require mode 1 or mode 3 groups")

    weighted: WeightedOptions | None = None
    if state.weighted_enabled:
        if not state.standard_deviations_path.strip():
            raise ValueError("a standard-deviation matrix is required for weighting")
        weighted = WeightedOptions(
            standard_deviations=load_matrix(state.standard_deviations_path),
            convergence_limit=state.weighted_convergence_limit,
            max_iterations=state.weighted_max_iterations,
        )

    kinetic: KineticOptions | None = None
    if state.kinetic_enabled:
        if not state.kinetic_models:
            raise ValueError("at least one kinetic model is required")
        kinetic = KineticOptions(
            models=tuple(
                kinetic_model_from_input(model) for model in state.kinetic_models
            )
        )

    return MCRALSOptions(
        max_iterations=state.max_iterations,
        tolerance=state.tolerance,
        divergence_limit=state.divergence_limit,
        normalization=state.normalization,  # type: ignore[arg-type]
        nonnegativity_c=NonnegativityOptions(
            enabled=state.nonnegative_c_enabled,
            algorithm=state.nonnegative_c_algorithm,  # type: ignore[arg-type]
            mask=parse_optional_mask(
                state.nonnegative_c_mask,
                n_components,
                name="concentration nonnegativity mask",
            ),
        ),
        nonnegativity_s=NonnegativityOptions(
            enabled=state.nonnegative_s_enabled,
            algorithm=state.nonnegative_s_algorithm,  # type: ignore[arg-type]
            mask=parse_optional_mask(
                state.nonnegative_s_mask,
                n_components,
                name="spectral nonnegativity mask",
            ),
        ),
        unimodality_c=UnimodalityOptions(
            enabled=state.unimodality_c_enabled,
            tolerance=state.unimodality_c_tolerance,
            mode=state.unimodality_c_mode,  # type: ignore[arg-type]
            mask=parse_optional_mask(
                state.unimodality_c_mask,
                n_components,
                name="concentration unimodality mask",
            ),
        ),
        unimodality_s=UnimodalityOptions(
            enabled=state.unimodality_s_enabled,
            tolerance=state.unimodality_s_tolerance,
            mode=state.unimodality_s_mode,  # type: ignore[arg-type]
            mask=parse_optional_mask(
                state.unimodality_s_mask,
                n_components,
                name="spectral unimodality mask",
            ),
        ),
        closure=ClosureOptions(
            enabled=state.closure_enabled,
            mode=state.closure_mode,  # type: ignore[arg-type]
            blocks=closure_blocks,
        ),
        concentration_values=_value_constraint(
            state.concentration_values_path, state.concentration_values_kind
        ),
        spectral_values=_value_constraint(
            state.spectral_values_path, state.spectral_values_kind
        ),
        multi=multi,
        trilinearity=TrilinearityOptions(
            enabled=state.trilinearity_enabled,
            direction=state.trilinearity_direction,  # type: ignore[arg-type]
            shape=state.trilinearity_shape,  # type: ignore[arg-type]
            component_mask=parse_optional_mask(
                state.trilinearity_component_mask,
                n_components,
                name="trilinearity component mask",
            ),
            quadrilinear_dimensions=quadrilinear,
        ),
        correlation=correlation,
        kinetic=kinetic,
        weighted=weighted,
        tucker=(
            TuckerOptions(
                enabled=True,
                n_matrices=(
                    None if state.tucker_n_matrices <= 0 else state.tucker_n_matrices
                ),
                modes=tuple(tucker_modes),
            )
            if state.tucker_enabled
            else None
        ),
    )


def create_initial_estimate(state: GUIState, data: ArrayLike) -> FloatArray:
    matrix = np.asarray(data, dtype=np.float64)
    method = state.initial_method.lower()
    if method == "file":
        if not state.initial_path.strip():
            raise ValueError("an initial-estimate file is required")
        return load_matrix(state.initial_path)
    if method == "simplisma":
        estimate, _ = pure(matrix, state.components, state.pure_noise_percent)
        return estimate
    if method == "efa":
        analysis = efa(matrix, n_factors=state.components)
        assert analysis.profiles is not None
        return analysis.profiles
    raise ValueError("initial method must be file, simplisma, or efa")


def state_to_dict(state: GUIState) -> dict[str, Any]:
    return asdict(state)


def state_from_dict(raw: Mapping[str, Any]) -> GUIState:
    allowed = {item.name for item in fields(GUIState)}
    values = {key: value for key, value in raw.items() if key in allowed}
    models = values.get("kinetic_models", [])
    values["kinetic_models"] = [
        (
            model
            if isinstance(model, KineticModelInput)
            else KineticModelInput(**dict(model))
        )
        for model in models
    ]
    return GUIState(**values)
